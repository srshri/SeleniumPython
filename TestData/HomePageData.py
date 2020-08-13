import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Shishir", "lastname": "Shrivastava","gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        workbook = openpyxl.load_workbook("D:\\Career\\Selenium(Python)\\PythonDemo.xlsx")
        sheet = workbook.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]